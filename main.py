from openpyxl import load_workbook
from openpyxl.utils import *

wb = load_workbook("textspel.xlsx")
ws = wb.active
room_number = 1
room_number_change_n = -1
room_number_change_s = 1
room_position = 'C'
start = True
wsRooms = wb["Rooms"]
wsMap = wb["map"]




def look():
    print(ws[F"{room_position}{room_number}"].value)


def north():
    global room_number
    if room_number >= 1:
        room_number += room_number_change_n
        print(ws[F"{room_position}{room_number}"].value)

def south():
    global room_number
    if room_number >= 1:
        room_number += room_number_change_s
        print(ws[F"{room_position}{room_number}"].value)

def west():
    global room_position
    if room_position == 1:
        room_position -= 1
        print(ws[F"{room_position}{room_number}"].value)



while start:
    reading = input("")
    if reading == "look":
        look()
    if reading == 'north':
        north()
    if reading == "south":
        south()
    elif reading == "stop":
        start = False
    if reading == 'quit':
        start = False




